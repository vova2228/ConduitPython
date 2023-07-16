import random
import string
import requests
from openpyxl import load_workbook
import sys

letters = string.ascii_lowercase
file_path = "tests/data/registered_users.xlsx"
register_endpoint = "https://api.realworld.io/api/users"
login_endpoint = "https://api.realworld.io/api/users/login"

bios = [
    "I love to code in Python!",
    "Software engineer and coffee addict",
    "Traveller and bookworm",
    "Foodie. Yoga enthusiast. Dog mom.",
    "Lifelong learner and problem solver"
]


def initialize_sheet():
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    return sheet


def insert_new_user_to_file(users_count: int):
    sheet = initialize_sheet()

    filled_rows = sheet.max_row

    if filled_rows == 1:
        i = 1
        while i <= users_count:
            random_length = random.randint(6, 11)
            add_user_to_sheet(sheet, i, random_length)
            i += 1
    else:
        i = filled_rows + 1
        while i <= filled_rows + users_count:
            random_length = random.randint(6, 11)
            add_user_to_sheet(sheet, i, random_length)
            i += 1

    sheet.parent.save(file_path)


def generate_email(length):
    username = ''.join(random.choice(letters) for i in range(length))
    domain = ''.join(random.choice(letters) for i in range(length))
    email = f"{username}@{domain}.com"
    return email


def generate_password(length):
    characters = string.ascii_letters + string.digits
    password = ''.join(random.choice(characters) for i in range(length))
    return password


def generate_username(length):
    username = ''.join(random.choice(letters) for i in range(length))
    return username


def generate_bio(length=20):
    bio = random.choice(bios)
    if len(bio) < length:
        extra = "".join(random.choices(string.ascii_lowercase, k=length - len(bio)))
        bio += " " + extra
    elif len(bio) > length:
        bio = bio[:length]

    return bio


def add_user_to_sheet(sheet, index, length):
    sheet[f'A{index}'] = generate_email(length)
    sheet[f'B{index}'] = generate_password(length)
    sheet[f'C{index}'] = generate_username(length)
    sheet[f'D{index}'] = generate_bio(length)


def register_users_from_file():
    sheet = initialize_sheet()
    filled_rows = sheet.max_row

    registered_users = 0
    index = 1
    while index <= filled_rows:
        email = sheet[f'A{index}'].value
        password = sheet[f'B{index}'].value
        username = sheet[f'C{index}'].value
        bio = sheet[f'D{index}'].value
        request_body = {
            "user": {
                "email": email,
                "password": password,
                "username": username,
                "bio": bio
            }
        }
        response = requests.post(register_endpoint, json=request_body)
        index += 1
        if "token" in response.text:
            registered_users += 1
    return registered_users


def clear_file():
    sheet = initialize_sheet()
    sheet.delete_rows(1, sheet.max_row)
    sheet.parent.save(file_path)


def main():
    clear_file()
    print("File cleared")
    number_of_user = sys.argv[1]
    insert_new_user_to_file(int(number_of_user))
    print("Users added")
    registered_users = register_users_from_file()
    print("Users registered. Number of registered users = ", registered_users)


if __name__ == "__main__":
    main()
