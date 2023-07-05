import random

from openpyxl import load_workbook


class Utils:

    @staticmethod
    def read_user_data(file_path='C:/Users/vorudyh/Desktop/Progs/Conduit_Pytest/data/registered_users.xlsx'):
        # Загрузка книги экселевского файла
        workbook = load_workbook(filename=file_path)
        # Выбор первого листа
        sheet = workbook.active
        filled_rows = 0
        for row in sheet.iter_rows():
            if any(cell.value for cell in row):
                filled_rows += 1

        index = random.randint(1, filled_rows)

        email = sheet[f'A{index}'].value
        password = sheet[f'B{index}'].value
        username = sheet[f'C{index}'].value

        return email, password, username

