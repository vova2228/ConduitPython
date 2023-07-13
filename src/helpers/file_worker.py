import random
from openpyxl import load_workbook
from src.generators.generator import TestDataGenerator


class FileWorker:

    """
    Handles reading and writing data to Excel files.

    Attributes:
        __users_file_path (str): Path to registered users Excel file.
        __articles_file_path (str): Path to articles Excel file.
    """

    __users_file_path = 'C:/Users/narut/Desktop/Idea/ConduitPython/tests/data/registered_users.xlsx'
    __articles_file_path = 'C:/Users/narut/Desktop/Idea/ConduitPython/tests/data/articles.xlsx'

    @classmethod
    def insert_new_user_to_file(cls, users_count: int):
        sheet = cls.initialize_users_sheet()

        filled_rows = sheet.max_row

        if filled_rows == 1:
            i = 1
            while i < users_count + 1:
                random_length = random.randint(6, 11)
                cls.__add_user_to_sheet(sheet, i, random_length)
                i += 1
        else:
            users_count += filled_rows
            i = filled_rows + 1
            while i <= users_count:
                random_length = random.randint(6, 11)
                cls.__add_user_to_sheet(sheet, i, random_length)
                i += 1

        sheet.parent.save(cls.__users_file_path)

    @classmethod
    def __add_user_to_sheet(cls, sheet, index, random_length):
        sheet[f'A{index}'] = TestDataGenerator.generate_email(random_length)
        sheet[f'B{index}'] = TestDataGenerator.generate_password(random_length)
        sheet[f'C{index}'] = TestDataGenerator.generate_username(random_length)

    @classmethod
    def initialize_users_sheet(cls):
        workbook = load_workbook(filename=cls.__users_file_path)
        sheet = workbook.active
        return sheet

    @classmethod
    def initialize_articles_sheet(cls):
        workbook = load_workbook(filename=cls.__articles_file_path)
        sheet = workbook.active
        return sheet

    @classmethod
    def get_user_from_file(cls):
        sheet = cls.initialize_users_sheet()
        filled_rows = sheet.max_row

        if filled_rows > 1:
            index = random.randint(1, filled_rows)
        else:
            index = 1

        email = sheet.cell(row=index, column=1).value
        password = sheet.cell(row=index, column=2).value
        username = sheet.cell(row=index, column=3).value

        return email, password, username

    @classmethod
    def insert_new_email_for_user(cls, old_email, new_email):
        sheet = cls.initialize_users_sheet()
        filled_rows = sheet.max_row

        for i in range(1, filled_rows + 1):
            current_email = sheet[f'A{i}'].value
            if current_email == old_email:
                sheet[f'A{i}'].value = new_email
                sheet.parent.save(cls.__users_file_path)
                return True
        return False

    @classmethod
    def get_article_from_file(cls):
        sheet = cls.initialize_articles_sheet()
        filled_rows = sheet.max_row

        index = random.randint(1, filled_rows)

        title = sheet.cell(row=index, column=1).value
        description = sheet.cell(row=index, column=2).value
        body = sheet.cell(row=index, column=3).value

        tags = [sheet.cell(row=i, column=4).value for i in range(index, filled_rows + 1) if
                sheet.cell(row=i, column=4).value]

        return title, description, body, tags
